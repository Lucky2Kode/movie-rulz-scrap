from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

MOVIE_HEADERS = ["Movie", "Year", "Lang", "Page", "Image File", "Trailer URL", "Date Added"]
DAILY_HEADERS = ["Date", "New Movies Added"]


def _load_existing_titles(path: Path, sheet_name: str = None) -> set[str]:
    if not path.exists():
        return set()
    wb = openpyxl.load_workbook(path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    titles = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            titles.add(str(row[0]).strip().lower())
    return titles


def filter_new(rows: list[dict], out_path: Path, sheet_name: str = None) -> list[dict]:
    """Return only rows whose movie title is not already in the given sheet."""
    existing = _load_existing_titles(out_path, sheet_name)
    return [r for r in rows if r.get("movie", "").strip().lower() not in existing]


def _autofit(ws):
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


def _rebuild_daily_sheet(wb):
    """Rebuild the Daily Summary sheet by aggregating across all data sheets."""
    if "Daily Summary" in wb.sheetnames:
        del wb["Daily Summary"]

    ds = wb.create_sheet("Daily Summary")
    ds.append(DAILY_HEADERS)
    for cell in ds[1]:
        cell.font = Font(bold=True)

    date_col = MOVIE_HEADERS.index("Date Added")
    counts: dict[str, int] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name == "Daily Summary":
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            day = str(row[date_col]).strip() if len(row) > date_col and row[date_col] else "unknown"
            counts[day] = counts.get(day, 0) + 1

    for day in sorted(counts):
        ds.append([day, counts[day]])

    _autofit(ds)


def append_new(rows: list[dict], out_path: Path, sheet_name: str = "Movies") -> list[dict]:
    """
    Append only movies not already in out_path.
    Creates a new tab if sheet_name doesn't exist yet (e.g. per-year bollywood tabs).
    Returns the list of newly added rows.
    Stamps each new row with today's date.
    Updates the Daily Summary tab automatically.
    """
    out_path.parent.mkdir(exist_ok=True)
    today = date.today().isoformat()

    new_rows = rows  # already filtered by caller via filter_new()

    if out_path.exists():
        wb = openpyxl.load_workbook(out_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(MOVIE_HEADERS)
            for cell in ws[1]:
                cell.font = Font(bold=True)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(MOVIE_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    # Insert new rows at the top (row 2, just below the header)
    ws.insert_rows(2, amount=len(new_rows))
    for i, row in enumerate(new_rows, start=2):
        ws.cell(i, 1, row.get("movie", ""))
        ws.cell(i, 2, row.get("year", ""))
        ws.cell(i, 3, row.get("lang", ""))
        ws.cell(i, 4, row.get("page", ""))
        ws.cell(i, 5, row.get("img_file", ""))
        ws.cell(i, 6, row.get("trailer_url", ""))
        ws.cell(i, 7, today)

    _autofit(ws)
    _rebuild_daily_sheet(wb)
    wb.save(out_path)

    print(f"\n{len(new_rows)} new movie(s) added to {out_path} [{sheet_name}]  [{today}]")
    return new_rows


def daily_summary(out_path: Path):
    """Print daily count to console."""
    if not out_path.exists():
        return
    wb = openpyxl.load_workbook(out_path)
    if "Daily Summary" not in wb.sheetnames:
        return
    ds = wb["Daily Summary"]
    print("\n--- Daily additions ---")
    for row in ds.iter_rows(min_row=2, values_only=True):
        if row[0]:
            print(f"  {row[0]}:  {row[1]} movie(s)")
